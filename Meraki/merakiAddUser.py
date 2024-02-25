import meraki
from decouple import config
import openpyxl
import warnings

API_KEY = config("API_KEY")
ORG_ID = config("ORG_ID")
TYPE = config("TYPE")

# This is here to remove just the warning about Excel validated cells
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# This method is called to get the data from the Excel
def getData():

    Excel_Doc = "add802_1X.xlsx"

    # Multiple tabs within the workbook can be called
    # It will return a list of dictionaries
    # Keys are the headers in the worksheet (Row 1 entries)
    # Values are the data from the rows (after the headers)

    wbn = Excel_Doc
    users = []
    # other = []

    # wb = Workbook being used
    wb = openpyxl.load_workbook(
        wbn,
        data_only=True,
    )

    # Gets the name of each tab in the Workbook
    tab_names = wb.sheetnames

    # Goes through each of the tabs to grab the data
    for tab in tab_names:
        ws = wb[tab]
        if tab == "WiFi_802.1X":
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                users.append(values)

        # elif tab == "Other":
        #     headers = [cell.value for cell in ws[1]]
        #     for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        #         # if row[0].value is not None:
        #         values = {}
        #         for key, cell in zip(headers, row):
        #             values[key] = cell.value
        #         other.append(values)

    return users


usersToAdd = getData()

dashboard = meraki.DashboardAPI(API_KEY, suppress_logging=True)

for user in usersToAdd:

    authorizations = [
        {
            "ssidNumber": int(user["ssidNumber"]),
            "expiresAt": user["expiresAt"],
        }  # Must use ISO 8601 Format
    ]

    dashboard.networks.createNetworkMerakiAuthUser(
        user["networkId"],
        user["email"],
        authorizations,
        name=user["name"],
        password=user["password"],
        accountType=TYPE,
        emailPasswordToUser=False,
        isAdmin=False,
    )
