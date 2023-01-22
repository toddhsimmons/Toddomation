import json
import openpyxl
import warnings 
from pprint import pprint

# This is here to remove just the warning about Excel validated cells
warnings.filterwarnings('ignore',
                        category=UserWarning, 
                        module='openpyxl'
                        )

def get_data():

    Excel_Doc = 'Toddomation_Meraki_Template.xlsx'
        # This method is called to get the data from the Excel
        # Multiple tabs within the workbook can be called
        # It will return a list of dictionaries 
        # Keys are the headers in the worksheet (Row 1 entries)
        # Values are the data from the rows (after the headers)

    wbn = Excel_Doc 
    org = []
    networks = []

    # wb = Workbook being used
    wb = openpyxl.load_workbook(
        wbn,
        data_only=True,
    )

    # Gets the name of each tab in the Workbook
    tab_names = wb.sheetnames

    # Goes through each of the tabs to grab the data
    for tab in tab_names:
        ws = wb[tab]
        if tab == 'Org':
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                org.append(values) 
        
        elif tab == 'Networks':
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                networks.append(values) 

    # print(org)
    # pprint(networks)
            
    return org,networks
