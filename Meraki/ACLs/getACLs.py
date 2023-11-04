import json
import openpyxl
import warnings
from icecream import ic

# This is here to remove just the warning about Excel validated cells
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def get_data(network_id):
    Excel_Doc = f"{network_id}.xlsx"
    # This method is called to get the data from the Excel
    # Multiple tabs within the workbook can be called
    # It will return a list of dictionaries
    # Keys are the headers in the worksheet (Row 1 entries)
    # Values are the data from the rows (after the headers)

    wbn = Excel_Doc
    switch = []
    firewallIn = []
    firewallOut = []
    wireless = []

    # wb = Workbook being used
    wb = openpyxl.load_workbook(
        wbn,
        data_only=True,
    )

    # Gets the name of each tab in the Workbook
    tab_names = wb.sheetnames

    # Goes through each of the tabs to grab the data
    for tab in tab_names:
        ws = wb[tab]
        if tab == "SwitchACL":
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                switch.append(values)

        elif tab == "FirewallL3InACL":
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                firewallIn.append(values)

        elif tab == "FirewallL3OutACL":
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                firewallOut.append(values)

        elif tab == "WirelessACL":
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # if row[0].value is not None:
                values = {}
                for key, cell in zip(headers, row):
                    values[key] = cell.value
                wireless.append(values)
    # print(org)
    # pprint(networks)

    return switch, firewallIn, firewallOut, wireless
